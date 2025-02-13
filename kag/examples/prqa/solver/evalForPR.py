import json

import pandas as pd
from kag.common.conf import KAG_CONFIG
from kag.solver.logic.solver_pipeline import SolverPipeline
from kag.common.registry import import_modules_from_path
from openpyxl import load_workbook
import os
import openpyxl

# def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):
#     try:
#         # Load existing workbook
#         book = load_workbook(filename)
#
#         # Create an ExcelWriter with the engine openpyxl
#         with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#             # This is now managed internally, and you do not need to set writer.book
#             writer.sheets = {ws.title: ws for ws in book.worksheets}
#
#             if not startrow and sheet_name in writer.sheets:
#                 startrow = writer.sheets[sheet_name].max_row
#
#             if startrow is None:
#                 startrow = 0
#
#             df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, header=not startrow, **to_excel_kwargs)
#
#     except FileNotFoundError:
#         # If file does not exist, this will create it
#         df.to_excel(filename, index=False, sheet_name=sheet_name, **to_excel_kwargs)

def qa(query):
    resp = SolverPipeline.from_config(KAG_CONFIG.all_config["kag_solver_pipeline"])
    answer, traceLog = resp.run(query)

    print(f"\n\nso the answer for '{query}' is: {answer}\n\n")  #
    print(traceLog)
    return answer, traceLog



if __name__ == "__main__":
    import_modules_from_path("./prompt")

    json_file_path = "./data/test.json"

    # with open(json_file_path, "r") as f:
    #     queries = json.load(f)

    # for item in queries:
    #     id = item.get("id")
    #     question = item.get("question")
    #     source = item.get("source")
    #     print(f"query {id}: {question}")
    #     answer, trace_log = qa(question)
    #     data = {
    #         'id': id,
    #         'query': question,
    #         "answer": answer,
    #         "trace_log": trace_log
    #     }
    #     with open("queries_output.txt", "a", encoding="utf-8") as file:
    #         # 将数据写入文件，每个字段一行
    #         file.write(f"id: {data['id']}\n")
    #         file.write(f"query: {data['query']}\n")
    #         file.write(f"answer: {data['answer']}\n")
    #         file.write(f"trace_log: {data['trace_log']}\n")
    #         file.write("\n")  # 每个记录之间添加一个空行



    query = "《酬韦相公见寄》的作者的某一个好友的好友张蠙除外, 请问他的好友还有谁?"
    answer, trace_log = qa(query)
    # print(f"Question: {query}")
    # print(f"Answer: {answer}")
    # print(f"TraceLog: {trace_log}")
